"""Extract Guess apparel images and write category / subcategory / color labels."""

from __future__ import annotations

import csv
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from label_map import map_category, map_color  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DATASET = ROOT / "dataset"
IMAGES = DATASET / "images"

GUESS_XLSX = ROOT / "exceli" / "6-ORDERSHEET_MAIN FW26_GUESS_WLCEE.xlsx"

NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
DRAWING_TYPE = f"{NS_R}/drawing"


def _local(tag: str) -> str:
    return tag.split("}")[-1]


def _attr_rid(attrib: dict) -> str | None:
    for key, value in attrib.items():
        if _local(key) == "id":
            return value
    return None


def _resolve(base: str, target: str) -> str:
    target = target.replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    parts = base.replace("\\", "/").split("/")[:-1]
    for chunk in target.split("/"):
        if chunk == "..":
            if parts:
                parts.pop()
        elif chunk not in ("", "."):
            parts.append(chunk)
    return "/".join(parts)


def _read_rels(zf: zipfile.ZipFile, rels_path: str) -> dict[str, dict]:
    root = ET.fromstring(zf.read(rels_path))
    rels = {}
    for rel in root:
        if _local(rel.tag) != "Relationship":
            continue
        rels[rel.attrib["Id"]] = {
            "target": rel.attrib["Target"],
            "type": rel.attrib.get("Type", ""),
        }
    return rels


def sheet_paths(zf: zipfile.ZipFile) -> dict[str, str]:
    wb_rels = _read_rels(zf, "xl/_rels/workbook.xml.rels")
    root = ET.fromstring(zf.read("xl/workbook.xml"))
    mapping = {}
    for el in root.iter():
        if _local(el.tag) != "sheet":
            continue
        name = el.attrib.get("name")
        rid = _attr_rid(el.attrib)
        if not name or not rid or rid not in wb_rels:
            continue
        mapping[name] = _resolve("xl/workbook.xml", wb_rels[rid]["target"])
    return mapping


def row_to_media(zf: zipfile.ZipFile, sheet_xml_path: str) -> dict[int, str]:
    rels_path = str(Path(sheet_xml_path).parent / "_rels" / (Path(sheet_xml_path).name + ".rels"))
    rels_path = rels_path.replace("\\", "/")
    if rels_path not in zf.namelist():
        return {}

    rels = _read_rels(zf, rels_path)
    drawing_path = None
    for rel in rels.values():
        if rel["type"] == DRAWING_TYPE or rel["type"].endswith("/drawing"):
            drawing_path = _resolve(sheet_xml_path, rel["target"])
            break
    if not drawing_path:
        return {}

    drawing_rels_path = str(
        Path(drawing_path).parent / "_rels" / (Path(drawing_path).name + ".rels")
    ).replace("\\", "/")
    drawing_rels = _read_rels(zf, drawing_rels_path) if drawing_rels_path in zf.namelist() else {}

    rid_to_media = {}
    for rid, rel in drawing_rels.items():
        media_path = _resolve(drawing_path, rel["target"])
        rid_to_media[rid] = media_path

    root = ET.fromstring(zf.read(drawing_path))
    row_media: dict[int, str] = {}
    for anchor in root.iter():
        if _local(anchor.tag) not in {"twoCellAnchor", "oneCellAnchor"}:
            continue
        row_el = None
        embed = None
        for child in anchor.iter():
            local = _local(child.tag)
            if local == "from":
                for marker in child:
                    if _local(marker.tag) == "row":
                        row_el = marker
            if local == "blip":
                embed = child.attrib.get(f"{{{NS_R}}}embed") or child.attrib.get("embed")
        if row_el is None or embed is None or embed not in rid_to_media:
            continue
        excel_row = int(row_el.text) + 1
        row_media[excel_row] = rid_to_media[embed]
    return row_media


def header_index(row) -> dict[str, int]:
    idx = {}
    for i, value in enumerate(row):
        if value is None or str(value).strip() == "":
            continue
        idx[str(value).strip()] = i
    return idx


def cell(row, idx: dict[str, int], name: str):
    if name not in idx:
        return None
    pos = idx[name]
    if pos >= len(row):
        return None
    return row[pos]


def safe_name(value) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value).strip())
    return text.strip("._") or "unknown"


def load_rules() -> tuple[dict, dict]:
    tree = json.loads((DATA / "item_tree_apparel.json").read_text(encoding="utf-8"))
    palette = json.loads((DATA / "color_palette.json").read_text(encoding="utf-8"))
    return tree, palette


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_summary(
    path: Path,
    *,
    source_file: str,
    n_rows: int,
    n_accessories: int,
    n_unique: int,
    labeled: list[dict],
    unmatched: list[dict],
) -> None:
    cat_counts = Counter(r["category"] for r in labeled)
    sub_counts = Counter(f"{r['category']} / {r['subcategory']}" for r in labeled)
    family_counts = Counter(r["color_family"] or "(unmapped)" for r in labeled)
    color_src = Counter(r["color_source"] for r in labeled)
    cat_src = Counter(r["category_source"] for r in labeled)
    reasons = Counter(r["reason"] for r in unmatched)
    confidence = Counter(r["match_confidence"] for r in labeled)

    lines = [
        "Guess apparel labeling summary",
        f"source_file: {source_file}",
        f"excel_data_rows: {n_rows}",
        f"skipped_accessories: {n_accessories}",
        f"unique_style_color_seen: {n_unique}",
        f"labeled: {len(labeled)}",
        f"unmatched: {len(unmatched)}",
        "",
        "match_confidence:",
    ]
    for key, n in confidence.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("category_source:")
    for key, n in cat_src.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("color_source:")
    for key, n in color_src.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("unmatched reasons:")
    for key, n in reasons.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("by category:")
    for key, n in cat_counts.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("by subcategory:")
    for key, n in sub_counts.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("by color_family:")
    for key, n in family_counts.most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("Not included yet: Levi's, Hugo Boss, Marciano, Kids, Footwear, Bags.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    tree, palette = load_rules()
    subcategory_to_category = tree["subcategory_to_category"]
    source_name = GUESS_XLSX.name

    IMAGES.mkdir(parents=True, exist_ok=True)

    labeled: list[dict] = []
    unmatched: list[dict] = []
    seen_skus: set[tuple[str, str]] = set()
    used_filenames: set[str] = set()
    n_rows = 0
    n_accessories = 0

    label_fields = [
        "image_path",
        "style",
        "color_code",
        "color_desc",
        "color_name",
        "color_family",
        "vendor_gh1",
        "part_desc",
        "category",
        "subcategory",
        "match_confidence",
        "source_file",
        "sheet",
        "gender",
        "category_source",
        "color_source",
    ]
    unmatched_fields = label_fields + ["reason"]

    with zipfile.ZipFile(GUESS_XLSX) as zf:
        sheets = sheet_paths(zf)
        print("sheets:", sheets)

        wb = load_workbook(GUESS_XLSX, data_only=True, read_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                media_by_row = row_to_media(zf, sheets[sheet_name])
                print(f"{sheet_name}: {len(media_by_row)} embedded images")

                rows = ws.iter_rows(values_only=True)
                header = next(rows)
                idx = header_index(header)

                for excel_row, row in enumerate(rows, start=2):
                    n_rows += 1
                    style = cell(row, idx, "Style")
                    color_code = cell(row, idx, "Color")
                    color_desc = cell(row, idx, "Color Desc")
                    gh1 = cell(row, idx, "GH1 Desc")
                    part_desc = cell(row, idx, "Part Desc")
                    gender = cell(row, idx, "Gender")

                    gh1_key = (str(gh1).strip().upper() if gh1 is not None else "")
                    if gh1_key == "ACCESSORIES":
                        n_accessories += 1
                        continue

                    style_s = "" if style is None else str(style).strip()
                    color_s = "" if color_code is None else str(color_code).strip()
                    if not style_s or not color_s:
                        unmatched.append(
                            {
                                "image_path": "",
                                "style": style_s,
                                "color_code": color_s,
                                "color_desc": color_desc,
                                "color_name": "",
                                "color_family": "",
                                "vendor_gh1": gh1,
                                "part_desc": part_desc,
                                "category": "",
                                "subcategory": "",
                                "match_confidence": "none",
                                "source_file": source_name,
                                "sheet": sheet_name,
                                "gender": gender,
                                "category_source": "",
                                "color_source": "",
                                "reason": "missing_style_or_color",
                            }
                        )
                        continue

                    sku = (style_s, color_s)
                    if sku in seen_skus:
                        continue
                    seen_skus.add(sku)

                    media_path = media_by_row.get(excel_row)
                    cat_match = map_category(gh1, part_desc, subcategory_to_category)
                    color_match = map_color(color_code, color_desc, palette)

                    if cat_match.source == "accessories":
                        n_accessories += 1
                        continue

                    has_image = bool(media_path and media_path in zf.namelist())
                    record = {
                        "image_path": "",
                        "style": style_s,
                        "color_code": color_s,
                        "color_desc": color_desc,
                        "color_name": color_match.color_name or "",
                        "color_family": color_match.color_family or "",
                        "vendor_gh1": gh1,
                        "part_desc": part_desc,
                        "category": cat_match.category or "",
                        "subcategory": cat_match.subcategory or "",
                        "match_confidence": "none",
                        "source_file": source_name,
                        "sheet": sheet_name,
                        "gender": gender,
                        "category_source": cat_match.source,
                        "color_source": color_match.source,
                    }

                    if not has_image:
                        record["reason"] = "missing_image"
                        unmatched.append(record)
                        continue
                    if not cat_match.category or not cat_match.subcategory:
                        record["reason"] = (
                            "unmapped_subcategory"
                            if cat_match.category
                            else "unmapped_category"
                        )
                        unmatched.append(record)
                        continue

                    filename = f"{safe_name(style_s)}_{safe_name(color_s)}.jpg"
                    if filename in used_filenames:
                        filename = f"{safe_name(style_s)}_{safe_name(color_s)}_{excel_row}.jpg"
                    used_filenames.add(filename)
                    dest = IMAGES / filename
                    dest.write_bytes(zf.read(media_path))
                    record["image_path"] = str(dest.relative_to(ROOT)).replace("\\", "/")

                    if cat_match.confidence == "high" and color_match.source == "exact_color_name":
                        record["match_confidence"] = "high"
                    elif cat_match.subcategory and color_match.color_family:
                        record["match_confidence"] = "medium"
                    elif cat_match.subcategory:
                        record["match_confidence"] = "low"
                    else:
                        record["match_confidence"] = "none"

                    labeled.append(record)
        finally:
            wb.close()

    write_csv(DATASET / "labels.csv", labeled, label_fields)
    write_csv(DATASET / "unmatched.csv", unmatched, unmatched_fields)
    write_summary(
        DATASET / "summary.txt",
        source_file=source_name,
        n_rows=n_rows,
        n_accessories=n_accessories,
        n_unique=len(seen_skus),
        labeled=labeled,
        unmatched=unmatched,
    )

    print(f"labeled={len(labeled)} unmatched={len(unmatched)} unique_sku={len(seen_skus)}")
    print(f"images dir: {IMAGES}")
    print(f"summary: {DATASET / 'summary.txt'}")


if __name__ == "__main__":
    main()
