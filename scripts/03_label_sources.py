"""Label Levi's, Hugo Boss, Marciano, Kids, Footwear and Bags, then merge into dataset."""

from __future__ import annotations

import csv
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from label_map import map_boss, map_category, map_color, map_levis  # noqa: E402
from xlsx_images import row_to_media, sheet_paths  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DATASET = ROOT / "dataset"
IMAGES = DATASET / "images"
EXCELI = ROOT / "exceli"

LABEL_FIELDS = [
    "image_path",
    "style",
    "color_code",
    "color_desc",
    "color_name",
    "color_family",
    "vendor_gh1",
    "part_desc",
    "category",
    "subcategory",
    "match_confidence",
    "source_file",
    "sheet",
    "gender",
    "category_source",
    "color_source",
    "division",
]
UNMATCHED_FIELDS = LABEL_FIELDS + ["reason"]

SOURCES = [
    {
        "tag": "guess",
        "file": "6-ORDERSHEET_MAIN FW26_GUESS_WLCEE.xlsx",
        "kind": "guess",
        "only_new": True,
    },
    {
        "tag": "marciano",
        "file": "8-ORDERSHEET_MAIN FW26_MARCIANO_WLCEE.xlsx",
        "kind": "guess",
    },
    {
        "tag": "kids",
        "file": "11-ORDERSHEET_MAINFW26_KIDS_WLCEE.xlsx",
        "kind": "guess",
    },
    {
        "tag": "ftw",
        "file": "10-ORDERSHEET_MAINFW26_FOOTWEAR_WLCEE.xlsx",
        "kind": "guess",
    },
    {
        "tag": "hb",
        "file": "7-ORDERSHEET_MAIN FW26_HB_WLCEE.xlsx",
        "kind": "guess",
    },
    {
        "tag": "levis",
        "file": "Order Levis Belgrade Usce APPAREL FW25.xlsx",
        "kind": "levis",
    },
    {
        "tag": "hwr",
        "file": "HWR FA26 order template incl all selections.xlsx",
        "kind": "boss",
        "sheets": ["HWR FA2026 all product"],
    },
    {
        "tag": "bmg",
        "file": "BMG_FA26_FCO_XYZ BANJA LUKA.xlsx",
        "kind": "boss",
        "sheets": ["XYZ BMG FA26", "BANJA LUKA BMG FA26"],
    },
]


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def header_index(row) -> dict[str, int]:
    idx = {}
    for i, value in enumerate(row):
        if value is None or str(value).strip() == "":
            continue
        key = str(value).strip()
        if key not in idx:
            idx[key] = i
    return idx


def cell(row, idx: dict[str, int], *names):
    for name in names:
        if name in idx and idx[name] < len(row):
            return row[idx[name]]
    return None


def safe_name(value) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value).strip())
    return text.strip("._") or "unknown"


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def dedupe_rows(rows: list[dict]) -> list[dict]:
    # isti artikal se ponavlja po velicini (S/M/L), slika je ista
    # zato dedup po style+color (ne samo style, jer boje trebaju za color task)
    seen_keys = set()
    unique = []
    for rec in rows:
        key = (rec.get("source_file", ""), rec.get("style", ""), rec.get("color_code", ""))
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique.append(rec)
    return unique


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def division_of(category: str, subcategory: str, tree: dict) -> str:
    for pair in tree["pairs"]:
        if pair["category"] == category and pair["subcategory"] == subcategory:
            return pair.get("division", "")
    return ""


def confidence_of(cat_match, color_match) -> str:
    if cat_match.confidence == "high" and color_match.source in {
        "exact_color_name",
        "levis_color_family",
    }:
        return "high"
    if cat_match.subcategory and color_match.color_family:
        return "medium"
    if cat_match.subcategory:
        return "low"
    return "none"


def empty_record(source_file, sheet) -> dict:
    return {key: "" for key in UNMATCHED_FIELDS} | {
        "source_file": source_file,
        "sheet": sheet,
        "match_confidence": "none",
    }


def save_image(zf, media_path, dest: Path) -> bool:
    if not media_path or media_path not in zf.namelist():
        return False
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(zf.read(media_path))
    return True


def finalize_record(record, cat_match, color_match, tree) -> dict:
    record["color_name"] = color_match.color_name or ""
    record["color_family"] = color_match.color_family or ""
    record["category"] = cat_match.category or ""
    record["subcategory"] = cat_match.subcategory or ""
    record["category_source"] = cat_match.source
    record["color_source"] = color_match.source
    record["division"] = division_of(record["category"], record["subcategory"], tree)
    record["match_confidence"] = confidence_of(cat_match, color_match)
    return record


def process_guess_like(source, zf, tree, palette, seen, used_names):
    path = EXCELI / source["file"]
    sheets = sheet_paths(zf)
    wb = load_workbook(path, data_only=True, read_only=True)
    labeled = []
    unmatched = []
    stats = Counter()
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name not in sheets:
                continue
            ws = wb[sheet_name]
            media_by_row = row_to_media(zf, sheets[sheet_name])
            rows = ws.iter_rows(values_only=True)
            header = None
            header_row = 0
            for i, row in enumerate(rows, start=1):
                idx_try = header_index(row or [])
                if "Style" in idx_try and "Color" in idx_try:
                    header = row
                    header_row = i
                    break
            if header is None:
                print(f"  {sheet_name}: skipped (no Style/Color header)")
                continue
            idx = header_index(header)
            print(f"  {sheet_name}: {len(media_by_row)} images, header_row={header_row}")
            for excel_row, row in enumerate(rows, start=header_row + 1):
                stats["rows"] += 1
                style = _norm(cell(row, idx, "Style"))
                color = _norm(cell(row, idx, "Color"))
                sku = (source["file"], style, color)
                if source.get("only_new") and sku in seen:
                    stats["already_labeled"] += 1
                    continue
                if sku in seen:
                    continue
                seen.add(sku)

                rec = empty_record(source["file"], sheet_name)
                rec.update(
                    {
                        "style": style,
                        "color_code": color,
                        "color_desc": cell(row, idx, "Color Desc"),
                        "vendor_gh1": cell(row, idx, "GH1 Desc"),
                        "part_desc": cell(row, idx, "Part Desc"),
                        "gender": cell(row, idx, "Gender"),
                    }
                )
                if not style or not color:
                    rec["reason"] = "missing_style_or_color"
                    unmatched.append(rec)
                    continue

                cat_match = map_category(
                    rec["vendor_gh1"],
                    rec["part_desc"],
                    tree["subcategory_to_category"],
                    gh0=cell(row, idx, "GH0 Desc"),
                    gh2=cell(row, idx, "GH2 Desc"),
                )
                color_match = map_color(color, rec["color_desc"], palette)
                finalize_record(rec, cat_match, color_match, tree)

                media_path = media_by_row.get(excel_row)
                if not rec["category"] or not rec["subcategory"]:
                    rec["reason"] = (
                        "unmapped_subcategory" if rec["category"] else "unmapped_category"
                    )
                    unmatched.append(rec)
                    continue

                filename = f"{source['tag']}_{safe_name(style)}_{safe_name(color)}.jpg"
                if filename in used_names:
                    filename = f"{source['tag']}_{safe_name(style)}_{safe_name(color)}_{excel_row}.jpg"
                dest = IMAGES / filename
                if not save_image(zf, media_path, dest):
                    rec["reason"] = "missing_image"
                    unmatched.append(rec)
                    continue
                used_names.add(filename)
                rec["image_path"] = str(dest.relative_to(ROOT)).replace("\\", "/")
                labeled.append(rec)
    finally:
        wb.close()
    return labeled, unmatched, stats


def process_levis(source, zf, tree, palette, seen, used_names):
    path = EXCELI / source["file"]
    sheets = sheet_paths(zf)
    wb = load_workbook(path, data_only=True, read_only=True)
    labeled = []
    unmatched = []
    stats = Counter()
    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        media_by_row = row_to_media(zf, sheets[sheet_name])
        print(f"  {sheet_name}: {len(media_by_row)} images")
        rows = ws.iter_rows(values_only=True)
        header = None
        header_row = 0
        for i, row in enumerate(rows, start=1):
            names = {str(x).strip() for x in row if x}
            if "Option" in names and "Category" in names and "SubCategory" in names:
                header = row
                header_row = i
                break
        idx = header_index(header or [])
        for excel_row, row in enumerate(rows, start=header_row + 1):
            stats["rows"] += 1
            option = _norm(cell(row, idx, "Option"))
            family = cell(row, idx, "Color Family")
            fashion_name = row[20] if len(row) > 20 else None
            sku = (source["file"], option, _norm(fashion_name))
            if not option or sku in seen:
                continue
            seen.add(sku)
            rec = empty_record(source["file"], sheet_name)
            rec.update(
                {
                    "style": option,
                    "color_code": _norm(fashion_name),
                    "color_desc": family,
                    "vendor_gh1": cell(row, idx, "Category"),
                    "part_desc": cell(row, idx, "SubCategory"),
                    "gender": cell(row, idx, "Gender"),
                }
            )
            cat_match = map_levis(
                rec["vendor_gh1"], rec["part_desc"], tree["subcategory_to_category"]
            )
            color_match = map_color(fashion_name, fashion_name, palette)
            if not color_match.color_family:
                color_match = map_color(fashion_name, family, palette)
            finalize_record(rec, cat_match, color_match, tree)

            if not rec["category"] or not rec["subcategory"]:
                rec["reason"] = "unmapped_subcategory"
                unmatched.append(rec)
                continue
            filename = f"{source['tag']}_{safe_name(option)}.jpg"
            dest = IMAGES / filename
            if not save_image(zf, media_by_row.get(excel_row), dest):
                rec["reason"] = "missing_image"
                unmatched.append(rec)
                continue
            used_names.add(filename)
            rec["image_path"] = str(dest.relative_to(ROOT)).replace("\\", "/")
            labeled.append(rec)
    finally:
        wb.close()
    return labeled, unmatched, stats


def process_boss(source, zf, tree, palette, seen, used_names):
    path = EXCELI / source["file"]
    sheets = sheet_paths(zf)
    wb = load_workbook(path, data_only=True, read_only=True)
    labeled = []
    unmatched = []
    stats = Counter()
    wanted = set(source.get("sheets") or wb.sheetnames)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name not in wanted or sheet_name.startswith("index#"):
                continue
            if sheet_name not in sheets:
                continue
            ws = wb[sheet_name]
            media_by_row = row_to_media(zf, sheets[sheet_name])
            print(f"  {sheet_name}: {len(media_by_row)} images")
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            idx = header_index(header or [])
            if "Style No." not in idx:
                continue
            for excel_row, row in enumerate(rows, start=2):
                stats["rows"] += 1
                style = _norm(cell(row, idx, "Style No."))
                color = _norm(cell(row, idx, "Color"))
                if not style:
                    continue
                sku = (source["file"], style, color)
                if sku in seen:
                    continue
                seen.add(sku)
                rec = empty_record(source["file"], sheet_name)
                rec.update(
                    {
                        "style": style,
                        "color_code": color,
                        "color_desc": color,
                        "vendor_gh1": cell(row, idx, "MPG"),
                        "part_desc": cell(row, idx, "SPG") or cell(row, idx, "Form Name"),
                    }
                )
                cat_match = map_boss(
                    rec["vendor_gh1"], rec["part_desc"], tree["subcategory_to_category"]
                )
                color_match = map_color(color, color, palette)
                finalize_record(rec, cat_match, color_match, tree)
                if not rec["category"] or not rec["subcategory"]:
                    rec["reason"] = "unmapped_subcategory"
                    unmatched.append(rec)
                    continue
                filename = f"{source['tag']}_{safe_name(style)}_{safe_name(color)}.jpg"
                dest = IMAGES / filename
                if not save_image(zf, media_by_row.get(excel_row), dest):
                    rec["reason"] = "missing_image"
                    unmatched.append(rec)
                    continue
                used_names.add(filename)
                rec["image_path"] = str(dest.relative_to(ROOT)).replace("\\", "/")
                labeled.append(rec)
    finally:
        wb.close()
    return labeled, unmatched, stats


def write_summary(path: Path, labeled: list[dict], unmatched: list[dict], per_source: dict) -> None:
    lines = [
        "Labeling summary (Apparel + Footware + Accesories)",
        f"labeled: {len(labeled)}",
        f"unmatched: {len(unmatched)}",
        "",
        "by source:",
    ]
    source_counts = Counter(r.get("source_file") or "" for r in labeled)
    for name, n in source_counts.most_common():
        rows = per_source.get(name, {}).get("rows", "")
        extra = f" unmatched={sum(1 for r in unmatched if r.get('source_file')==name)}"
        if rows:
            extra += f" excel_rows={rows}"
        lines.append(f"  {name}: labeled={n}{extra}")
    lines += ["", "by division:"]
    for key, n in Counter(r.get("division") or "(none)" for r in labeled).most_common():
        lines.append(f"  {key}: {n}")
    lines += ["", "by category:"]
    for key, n in Counter(r["category"] for r in labeled).most_common():
        lines.append(f"  {key}: {n}")
    lines += ["", "by subcategory:"]
    for key, n in Counter(f"{r['category']} / {r['subcategory']}" for r in labeled).most_common():
        lines.append(f"  {key}: {n}")
    lines += ["", "by color_family:"]
    for key, n in Counter(r["color_family"] or "(unmapped)" for r in labeled).most_common():
        lines.append(f"  {key}: {n}")
    lines += ["", "unmatched reasons:"]
    for key, n in Counter(r.get("reason") or "" for r in unmatched).most_common():
        lines.append(f"  {key}: {n}")
    lines += ["", "unmatched by source:"]
    for key, n in Counter(r.get("source_file") or "" for r in unmatched).most_common():
        lines.append(f"  {key}: {n}")
    lines.append("")
    lines.append("Skipped as duplicate SKU sheets: Marciano order form, FTW order form, HB order form, HWR selection subsets.")
    lines.append("Skipped Premiata: order form is not a tabular product catalog.")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    tree_path = DATA / "item_tree.json"
    if not tree_path.exists():
        raise SystemExit("Run scripts/01_parse_rules.py first.")
    tree = load_json(tree_path)
    palette = load_json(DATA / "color_palette.json")

    existing = dedupe_rows(load_csv(DATASET / "labels.csv"))
    for rec in existing:
        rec.setdefault("division", "")
        rec.setdefault("category_source", rec.get("category_source", ""))
        rec.setdefault("color_source", rec.get("color_source", ""))
        if not rec.get("division") and rec.get("category") and rec.get("subcategory"):
            rec["division"] = division_of(rec["category"], rec["subcategory"], tree)
    labeled = list(existing)
    unmatched: list[dict] = []
    seen = {(r.get("source_file", ""), r.get("style", ""), r.get("color_code", "")) for r in existing}
    used_names = {Path(r["image_path"]).name for r in existing if r.get("image_path")}
    per_source = defaultdict(lambda: Counter())
    per_source["existing Guess apparel"]["labeled"] = len(existing)

    IMAGES.mkdir(parents=True, exist_ok=True)

    processors = {
        "guess": process_guess_like,
        "levis": process_levis,
        "boss": process_boss,
    }

    for source in SOURCES:
        path = EXCELI / source["file"]
        print(f"\n=== {source['file']} ===")
        with zipfile.ZipFile(path) as zf:
            new_labeled, new_unmatched, stats = processors[source["kind"]](
                source, zf, tree, palette, seen, used_names
            )
        labeled.extend(new_labeled)
        unmatched.extend(new_unmatched)
        per_source[source["file"]]["labeled"] = len(new_labeled)
        per_source[source["file"]]["unmatched"] = len(new_unmatched)
        per_source[source["file"]]["rows"] = stats["rows"]
        print(f"  added labeled={len(new_labeled)} unmatched={len(new_unmatched)}")

    labeled = dedupe_rows(labeled)
    unmatched = dedupe_rows(unmatched)
    write_csv(DATASET / "labels.csv", labeled, LABEL_FIELDS)
    write_csv(DATASET / "unmatched.csv", unmatched, UNMATCHED_FIELDS)
    write_summary(DATASET / "summary.txt", labeled, unmatched, per_source)
    print(f"\nTOTAL labeled={len(labeled)} unmatched={len(unmatched)}")
    print(DATASET / "summary.txt")


if __name__ == "__main__":
    main()
