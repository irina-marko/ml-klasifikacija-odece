"""Parse Item Tree (English Apparel, Footware, Accesories) and the FCO color palette.

Reads:
  pravila/Item Tree.xlsx          sheet Klasifikacija (Full L4/L3/L2)
  pravila/Paleta Fashion Company boja.xlsx  sheet Colors

Writes:
  data/item_tree.json
  data/item_tree.csv
  data/item_tree_apparel.json
  data/item_tree_apparel.csv
  data/color_palette.json
  data/color_palette.csv
"""

from __future__ import annotations

import csv
import json
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PRAVILA = ROOT / "pravila"
DATA = ROOT / "data"

ITEM_TREE_XLSX = PRAVILA / "Item Tree.xlsx"
PALETTE_XLSX = PRAVILA / "Paleta Fashion Company boja.xlsx"

# Folder-only rows in Klasifikacija where L4 == L3. Keep T-shirt and Dresses
# because they are also real product types (generic tee / generic dress).
ALLOWED_DIVISIONS = {"apparel", "footware", "accesories"}

# Folder-only rows in Klasifikacija where L4 == L3. Keep T-shirt, Dresses,
# Sneakers, Boots, Bags, Belts, etc. because they are also real product types.
SKIP_AS_PRODUCT_TYPE = {
    "apparel",
    "outerwear",
    "tops",
    "bottoms",
    "suits & sets",
    "sleepwear",
    "underwear",
    "swimwear",
    "footware",
    "accesories",
}


def _norm(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def parse_item_tree(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Klasifikacija"]

    pairs: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    categories: OrderedDict[str, None] = OrderedDict()
    divisions: OrderedDict[str, None] = OrderedDict()

    for row in ws.iter_rows(min_row=4, values_only=True):
        product_type = _norm(row[11] if len(row) > 11 else None)  # Full L4
        product_family = _norm(row[12] if len(row) > 12 else None)  # Full L3
        l2_category = _norm(row[13] if len(row) > 13 else None)  # Full L2

        if l2_category.lower() not in ALLOWED_DIVISIONS:
            continue
        if not product_family or not product_type:
            continue
        if product_family.lower() in {"apparel", "footware", "accesories"} and product_type.lower() == product_family.lower():
            continue

        divisions.setdefault(l2_category, None)
        categories.setdefault(product_family, None)

        if (
            product_type.lower() == product_family.lower()
            and product_type.lower() in SKIP_AS_PRODUCT_TYPE
        ):
            continue

        key = (l2_category, product_family, product_type)
        if key in seen:
            continue
        seen.add(key)
        pairs.append(
            {
                "division": l2_category,
                "category": product_family,
                "subcategory": product_type,
            }
        )

    wb.close()

    subcategory_to_category: OrderedDict[str, str] = OrderedDict()
    conflicts: list[dict] = []
    for pair in pairs:
        sub = pair["subcategory"]
        cat = pair["category"]
        if sub in subcategory_to_category and subcategory_to_category[sub] != cat:
            conflicts.append(
                {
                    "subcategory": sub,
                    "first_category": subcategory_to_category[sub],
                    "other_category": cat,
                }
            )
            continue
        subcategory_to_category[sub] = cat

    return {
        "source": str(xlsx_path.name),
        "language": "en",
        "divisions": list(divisions.keys()),
        "categories": list(categories.keys()),
        "pairs": pairs,
        "subcategory_to_category": subcategory_to_category,
        "conflicts": conflicts,
    }


def parse_palette(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Colors"]

    colors: list[dict] = []
    families: OrderedDict[str, None] = OrderedDict()
    name_to_family: OrderedDict[str, str] = OrderedDict()
    header_found = False

    for row in ws.iter_rows(values_only=True):
        code = row[0] if row else None
        if not header_found:
            if code == "Color Code":
                header_found = True
            continue
        if not code:
            continue

        color_name = _norm(row[1])
        family = _norm(row[2])
        family_code = row[3]
        hex_code = _norm(row[4])
        rgb = {
            "r": row[5],
            "g": row[6],
            "b": row[7],
        }
        colors.append(
            {
                "color_code": _norm(code),
                "color_name": color_name,
                "family": family,
                "family_code": family_code,
                "hex": hex_code,
                "rgb": rgb,
            }
        )
        if family:
            families.setdefault(family, None)
        if color_name:
            name_to_family[color_name] = family

    wb.close()

    return {
        "source": str(xlsx_path.name),
        "families": list(families.keys()),
        "colors": colors,
        "name_to_family": name_to_family,
    }


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def write_tree_csv(path: Path, tree: dict) -> None:
    fieldnames = ["division", "category", "subcategory"]
    if tree["pairs"] and "division" not in tree["pairs"][0]:
        fieldnames = ["category", "subcategory"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(tree["pairs"])


def write_palette_csv(path: Path, palette: dict) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["color_code", "color_name", "family", "family_code", "hex", "r", "g", "b"],
        )
        writer.writeheader()
        for color in palette["colors"]:
            writer.writerow(
                {
                    "color_code": color["color_code"],
                    "color_name": color["color_name"],
                    "family": color["family"],
                    "family_code": color["family_code"],
                    "hex": color["hex"],
                    "r": color["rgb"]["r"],
                    "g": color["rgb"]["g"],
                    "b": color["rgb"]["b"],
                }
            )


def main() -> None:
    DATA.mkdir(parents=True, exist_ok=True)

    tree = parse_item_tree(ITEM_TREE_XLSX)
    palette = parse_palette(PALETTE_XLSX)
    apparel_pairs = [p for p in tree["pairs"] if p.get("division") == "Apparel"]
    apparel_tree = {
        **tree,
        "division": "Apparel",
        "divisions": ["Apparel"],
        "categories": list(
            OrderedDict((p["category"], None) for p in apparel_pairs).keys()
        ),
        "pairs": [{"category": p["category"], "subcategory": p["subcategory"]} for p in apparel_pairs],
        "subcategory_to_category": {
            p["subcategory"]: p["category"] for p in apparel_pairs
        },
    }

    write_json(DATA / "item_tree.json", tree)
    write_tree_csv(DATA / "item_tree.csv", tree)
    write_json(DATA / "item_tree_apparel.json", apparel_tree)
    write_tree_csv(DATA / "item_tree_apparel.csv", apparel_tree)
    write_json(DATA / "color_palette.json", palette)
    write_palette_csv(DATA / "color_palette.csv", palette)

    print(f"Divisions: {', '.join(tree['divisions'])}")
    print(f"Categories ({len(tree['categories'])}):")
    for category in tree["categories"]:
        subs = [p["subcategory"] for p in tree["pairs"] if p["category"] == category]
        print(f"  {category}: {', '.join(subs)}")
    print(f"\nProduct type pairs: {len(tree['pairs'])}")
    if tree["conflicts"]:
        print("WARNING category conflicts:")
        for conflict in tree["conflicts"]:
            print(f"  {conflict}")
    else:
        print("No subcategory-to-category conflicts.")

    print(f"\nColor families ({len(palette['families'])}): {', '.join(palette['families'])}")
    print(f"Color names: {len(palette['colors'])}")
    print(f"\nWrote files in {DATA}")


if __name__ == "__main__":
    main()
